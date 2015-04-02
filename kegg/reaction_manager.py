import urllib2

import openpyxl

from serialization.xlsx_helper import HEADER_STYLE, add_values, get_info


KEGG_REACTION_FILE = "/home/anna/Documents/IBGC/Models/KEGG_reactions.xlsx"
KEGG_COMPOUND_FILE = "/home/anna/Documents/IBGC/Models/KEGG_compounds.xlsx"


def get_rns_by_elements(elements):
    rns = set()
    reactions_by_compounds = urllib2.urlopen(
        'http://rest.kegg.jp/link/reaction/%s' % "+".join(elements)).read()
    for line in reactions_by_compounds.split("\n"):
        if line.find("rn:") != -1:
            _, rn = line.split("\t")
            rns.add(rn)
    return rns


def get_compounds_by_rn(rn):
    compounds_by_reaction = urllib2.urlopen('http://rest.kegg.jp/link/compound/%s' % rn).read()
    return {c for c in {c.replace("%s\t" % rn, '').strip() for c in compounds_by_reaction.split("\n")} if
            c.find('cpd:') != -1}


def get_compounds_by_rp(rp):
    compounds_by_rp = urllib2.urlopen('http://rest.kegg.jp/link/compound/%s' % rp).read()
    return {c for c in {c.replace("%s\t" % rp, '').strip() for c in compounds_by_rp.split("\n")} if
            c.find('cpd:') != -1}


def get_rpairs_by_rn(rn):
    rpairs_by_reaction = urllib2.urlopen('http://rest.kegg.jp/link/rpair/%s' % rn).read()
    return {c for c in {c.replace("%s\t" % rn, '').strip() for c in rpairs_by_reaction.split("\n")} if
            c.find('rp:') != -1}


def get_rpairs_by_compounds(compounds):
    rpairs_by_compounds = urllib2.urlopen('http://rest.kegg.jp/link/rpair/%s' % '+'.join(compounds)).read()
    rpairs = set()
    for line in rpairs_by_compounds.split("\n"):
        if line.find("rp:") != -1:
            _, rp = line.split("\t")
            rpairs.add(rp)
    return rpairs


def get_pw_name(org, pw):
    result = urllib2.urlopen('http://rest.kegg.jp/find/pathway/%s' % p_id_specific2generic(org, pw)).read()
    result = result.replace("\n", '')
    pw, pw_name = result.split('\t')
    return pw_name


def get_rns():
    rn2formula = {}
    reactions = urllib2.urlopen('http://rest.kegg.jp/list/reaction').read()
    for line in reactions.split("\n"):
        if line.find("rn:") != -1:
            rn, r_formula = line.split("\t")
            rn2formula[rn] = r_formula
    return rn2formula


def get_kegg_r_info(rn):
    name, definition, equation, ec, orthology = '', '', '', '', ''
    try:
        reactions = urllib2.urlopen('http://rest.kegg.jp/get/%s' % rn).read()
        for line in reactions.split("\n"):
            if line.find("DEFINITION") != -1:
                definition = line.replace("DEFINITION", '').strip()
            elif line.find("NAME") != -1:
                name = line.replace("NAME", '').strip()
            elif line.find("EQUATION") != -1:
                equation = line.replace("EQUATION", '').strip()
            elif line.find("ENZYME") != -1:
                ec = line.replace("ENZYME", '').strip()
            elif line.find("ORTHOLOGY") != -1:
                orthology = line.replace("ORTHOLOGY", '').strip()
    except:
        pass
    return name, definition, equation, ec, orthology


def get_kegg_c_info(cpd):
    names, formula = {}, ''
    try:
        compound = urllib2.urlopen('http://rest.kegg.jp/get/%s' % cpd).read()
        cur_line = ''
        for line in compound.split("\n"):
            if line.find(';') != -1:
                cur_line += line
                continue
            line = cur_line + line
            cur_line = ''
            if line.find("NAME") != -1:
                names = {it.strip() for it in line.replace("NAME", '').strip().split(';') if it.strip()}
            elif line.find("FORMULA") != -1:
                formula = line.replace("FORMULA", '').strip()
            if names and formula:
                return names, formula
    except:
        pass
    return names, formula


def serialize_reactions(path=KEGG_REACTION_FILE):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(0, "Reactions")
    add_values(ws, 1, 1, ["Id", "Name", "Definition", "Formula", "EC", "Orthology"], HEADER_STYLE)
    row = 2
    rns = {it.split('\t')[0] for it in urllib2.urlopen('http://rest.kegg.jp/list/reaction').read().split("\n") if
             it.find('rn:') != -1}
    for rn in rns:
        name, definition, equation, ec, orthology = get_kegg_r_info(rn)
        add_values(ws, row, 1, [rn.replace('rn:', '').strip(), name, definition, equation, ec, orthology])
        row += 1
    wb.save(path)


def serialize_compounds(path=KEGG_COMPOUND_FILE):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(0, "Compounds")
    add_values(ws, 1, 1, ["Id", "Name", "Formula"], HEADER_STYLE)
    row = 2
    cpds = {it.split('\t')[0] for it in urllib2.urlopen('http://rest.kegg.jp/list/compound').read().split("\n") if
             it.find('cpd:') != -1}
    for cpd in cpds:
        names, formula = get_kegg_c_info(cpd)
        add_values(ws, row, 1, [cpd.replace('cpd:', '').strip(), '; '.join(names), formula])
        row += 1
    wb.save(path)


def get_formula2kegg_compound(path=KEGG_COMPOUND_FILE):
    formula2kegg = {}
    # there are following columns: "Id", "Name", "Formula",
    # but we are only interested in kegg id (1) and formula (3)
    for (kegg, formula) in get_info(path, [1, 3], 2):
        if not formula:
            continue
        kegg, formula = str(kegg), str(formula)
        if formula:
            formula2kegg[formula] = kegg
    return formula2kegg


def get_rs_ps_by_kegg_equation(equation, stoichiometry=False):
    # Equation looks somewhat like '2 C00430 <=> C00931 + 2 C00001'
    rs, ps = equation.split(' <=> ')
    rs, ps = rs.split(' + '), ps.split(' + ')
    if stoichiometry:
        def val2pair(val):
            st_m = val.split(' ')
            (st, m) = (1, st_m[0]) if len(st_m) == 1 else st_m
            try:
                st = float(st)
            except ValueError:
                st = 1
            return m, st
        rs = {val2pair(m) for m in rs}
        ps = {val2pair(m) for m in ps}
    else:
        rs = {m[0] if len(m) == 1 else m[1] for m in (m.split(' ') for m in rs)}
        ps = {m[0] if len(m) == 1 else m[1] for m in (m.split(' ') for m in ps)}
    return rs, ps


def get_compounds2rn(path=KEGG_REACTION_FILE):
    rs_ps2rn = {}
    # there are following columns: "Id", "Name", "Definition", "Formula", "EC", "Orthology",
    # but we are only interested in kegg id (1) and formula (4)
    for (kegg, formula) in get_info(path, [1, 4], 2):
        if not formula:
            continue
        kegg, formula = str(kegg), str(formula)
        rs, ps = get_rs_ps_by_kegg_equation(formula)
        rs, ps = tuple(sorted(rs)), tuple(sorted(ps))
        if rs > ps:
            rs, ps = ps, rs
        rs_ps2rn[(rs, ps)] = kegg
    return rs_ps2rn


def get_rn2kegg_formula(path=KEGG_REACTION_FILE):
    rn2formula = {}
    # there are following columns: "Id", "Name", "Definition", "Formula", "EC", "Orthology",
    # but we are only interested in kegg id(1) and definition (3), i.e. a human-readable formula
    for (kegg, formula) in get_info(path, [1, 3], 2):
        if not formula:
            continue
        rn2formula[str(kegg)] = str(formula)
    return rn2formula


def get_rn2compounds(path=KEGG_REACTION_FILE, stoichiometry=False):
    rn2rs_ps = {}
    # there are following columns: "Id", "Name", "Definition", "Formula", "EC", "Orthology",
    # but we are only interested in kegg id (1) and formula (4)
    for (kegg, formula) in get_info(path, [1, 4], 2):
        if not formula:
            continue
        kegg, formula = str(kegg), str(formula)
        rn2rs_ps[kegg] = get_rs_ps_by_kegg_equation(formula, stoichiometry)
    return rn2rs_ps


def get_rn2name(path=KEGG_REACTION_FILE):
    rn2name = {}
    # there are following columns: "Id", "Name", "Definition", "Formula", "EC", "Orthology",
    # but we are only interested in kegg id (1) and name (2)
    for (kegg, name) in get_info(path, [1, 2], 2):
        rn2name[str(kegg)] = str(name)
    return rn2name


# serialize_compounds()