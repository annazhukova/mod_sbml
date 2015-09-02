from collections import defaultdict
import logging
import re

import libsbml
import openpyxl

from mod_cobra.gibbs.reaction_boundary_manager import get_bounds
from mod_sbml.sbml.sbml_manager import get_gene_association, get_reactants, get_products
from mod_sbml.serialization.xlsx_helper import save_data, add_values, HEADER_STYLE, BASIC_STYLE, RED_STYLE
from mod_sbml.annotation.kegg.kegg_annotator import get_kegg_m_id, get_kegg_r_id

__author__ = 'anna'


def serialize_model_info(sbml, path, r_style=lambda r_id: BASIC_STYLE):
    doc = libsbml.SBMLReader().readSBML(sbml)
    model = doc.getModel()
    wb = openpyxl.Workbook()
    save_data(["Id", "Name", "Compartment", "Kegg"], (
        (m.id, m.name, model.getCompartment(m.getCompartment()).name, get_kegg_m_id(m))
        for m in sorted(model.getListOfSpecies(), key=lambda m: m.id)), ws_index=0, ws_name="Metabolites", wb=wb)
    save_data(["Id", "Name", "Lower Bound", "Upper Bound", "Formula", "Kegg", "Gene association"],
              ((r.id, r.name, get_bounds(r)[0], get_bounds(r)[1], get_sbml_r_formula(model, r, False), get_kegg_r_id(r),
                get_gene_association(r))
               for r in sorted(model.getListOfReactions(), key=lambda r: r.id)),
              styles=(r_style(r.id) for r in sorted(model.getListOfReactions(), key=lambda r: r.id)),
              ws_index=1, ws_name="Reactions", wb=wb)
    wb.save(path)


def read_reactions(path):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.get_active_sheet()
    metabolites = defaultdict(set)
    compartments = set()
    r_name2info = {}
    for row in worksheet.rows:
        r_name = str(row[1].value).encode('utf-8')
        r_formula = str(row[2].value).encode('utf-8')
        reversible = str(row[3].value).encode('utf-8').lower().find("irreversible") == -1
        reactants, products = r_formula.split(" = ")
        reactants = [r.replace('"', '').strip() for r in reactants.split(" + ")] if reactants.strip() else []
        products = [r.replace('"', '').strip() for r in products.split(" + ")] if products.strip() else []
        r2stoich = get_stoichiometries(reactants, metabolites, compartments)
        p2stoich = get_stoichiometries(products, metabolites, compartments)
        r_name2info[r_name] = r2stoich, p2stoich, reversible
    return metabolites, compartments, r_name2info


def get_stoichiometries(participants, metabolites, compartments):
    m2stoich = {}
    for m in participants:
        if m.find(" * ") != -1:
            st, m = m.split(" * ")
            st = int(st)
        else:
            st = 1
        comp = None
        comps = re.findall(r' \[\w+\]', m)
        for comp in comps:
            m = m.replace(comp, "")
            comp = comp[2:-1].strip()
            compartments.add(comp)
            break
        metabolites[m].add(comp)
        m2stoich[m] = st, comp
    return m2stoich


def get_sbml_r_formula(model, r, show_compartments=True, show_metabolite_ids=True):
    format_m = lambda m_id, st: "%s%s" % (
        "%g " % st if st != 1 else "",
        format_m_name(model.getSpecies(m_id), model, show_compartments, show_metabolite_ids))
    formula = " + ".join([format_m(m_id, st) for (m_id, st) in get_reactants(r, True)]) + \
              (" <=> " if r.getReversible() else "-->") + \
              " + ".join([format_m(m_id, st) for (m_id, st) in get_products(r, True)])
    return formula


def format_m_name(m, model, show_compartment=True, show_id=True):
    name = m.getName()
    c = model.getCompartment(m.getCompartment())
    c_name = c.getName()
    if -1 == name.find(c_name) and show_compartment:
        name = "%s[%s]" % (name, c_name)
    if not show_compartment:
        name = name.replace("[%s]" % c_name, "").replace("[%s]" % c.getId(), "").strip()
    return "%s(%s)" % (name, m.id) if show_id else name


def get_cobra_r_formula(r, comp=True):
    if comp:
        return "{0} <=> {1}".format(
            " + ".join(
                ["%s%s[%s](%s)" % (('%g ' % -r.get_coefficient(m)) if r.get_coefficient(m) != -1 else '',
                                   m.name, m.compartment, m.id) for m in r.reactants]),
            " + ".join(["%s%s[%s](%s)" % (('%g ' % r.get_coefficient(m)) if r.get_coefficient(m) != 1 else '',
                                          m.name, m.compartment, m.id) for m in r.products]))
    else:
        return "{0} <=> {1}".format(
            " + ".join(
                ["%s%s(%s)" % (('%g ' % -r.get_coefficient(m)) if r.get_coefficient(m) != -1 else '',
                               m.name, m.id) for m in r.reactants]),
            " + ".join(["%s%s(%s)" % (('%g ' % r.get_coefficient(m)) if r.get_coefficient(m) != 1 else '',
                                      m.name, m.id) for m in r.products]))


def save_model_mappings(source_model, target_model,
                        (c_id_s2t, c_ids_unmapped,\
                         m_id2m_id, m_id2m_id_diff_comp, m_ids_unmapped, \
                         r_id2r_id, r_id2r_id_diff_comp, r_ids_unmapped, \
                         s_s_id2chebi_id, t_s_id2chebi_id),
                        filename):
    logging.info('==========================\nCompartments: %d, %d\nMetabolites: %d %d %d\nReactions: %d %d %d' %
                 (len(c_id_s2t), len(c_ids_unmapped),
                  len(m_id2m_id), len(m_id2m_id_diff_comp), len(m_ids_unmapped),
                  len(r_id2r_id), len(r_id2r_id_diff_comp), len(r_ids_unmapped)))

    workbook = openpyxl.Workbook()
    c_ws = workbook.create_sheet(0, "Compartment mapping")
    save_compartment_mappings(c_ws, source_model, target_model, c_id_s2t, c_ids_unmapped)
    m_ws = workbook.create_sheet(1, "Metabolite mapping")
    save_metabolite_mappings(m_ws, source_model, target_model,
                             m_id2m_id, m_id2m_id_diff_comp, m_ids_unmapped, get_kegg_m_id,
                             lambda m_id, model: (
                                 s_s_id2chebi_id[m_id] if m_id in s_s_id2chebi_id else None) if model == source_model
                             else (
                                 t_s_id2chebi_id[m_id] if m_id in t_s_id2chebi_id else None))
    r_ws = workbook.create_sheet(2, "Reaction mapping")
    save_reaction_mappings(r_ws, source_model, target_model,
                           r_id2r_id, r_id2r_id_diff_comp, r_ids_unmapped, get_kegg_r_id)
    workbook.save(filename)


def save_metabolite_mappings(ws, source_model, target_model, m_id2m_id, m_id2m_id_diff_comp,
                             s_m_ids_unmapped, m2kegg, m2chebi):
    headers = ["Name", "Id", "Compartment", "ChEBI", "KEGG"]
    add_values(ws, 1, 2, [source_model.name if source_model.name else source_model.id], HEADER_STYLE)
    add_values(ws, 2, 2, reversed(headers), HEADER_STYLE)
    add_values(ws, 1, 2 + len(headers), [target_model.name if target_model.name else target_model.id], HEADER_STYLE)
    add_values(ws, 2, 2 + len(headers), headers, HEADER_STYLE)

    ws.cell(row=3, column=1).value = "Matching"
    i = 4
    for s_m_id in sorted(m_id2m_id.iterkeys()):
        add_metabolite(ws, s_m_id, source_model, i, 2, m2kegg, m2chebi, rev=True)
        t_m_id = m_id2m_id[s_m_id]
        add_metabolite(ws, t_m_id, target_model, i, 2 + len(headers), m2kegg, m2chebi)
        i += 1

    ws.cell(row=i, column=1).value = "Similar in different compartments"
    i += 1
    for s_m_id in sorted(m_id2m_id_diff_comp.iterkeys()):
        add_metabolite(ws, s_m_id, source_model, i, 2, m2kegg, m2chebi, rev=True)
        for t_m_id in m_id2m_id_diff_comp[s_m_id]:
            add_metabolite(ws, t_m_id, target_model, i, 2 + len(headers), m2kegg, m2chebi)
            i += 1

    ws.cell(row=i, column=1).value = "Unmapped"
    i += 1
    for s_m_id in sorted(s_m_ids_unmapped):
        add_metabolite(ws, s_m_id, source_model, i, 2, m2kegg, m2chebi, rev=True)
        i += 1


def save_reaction_mappings(ws, source_model, target_model, r_id2r_id, r_id2r_id_diff_comp,
                           s_r_ids_unmapped, r2kegg):
    headers = ["Formula", "Id", "Name", "L. b.", "Up. b.", "KEGG"]
    add_values(ws, 1, 2, [source_model.name if source_model.name else source_model.id], HEADER_STYLE)
    add_values(ws, 2, 2, reversed(headers), HEADER_STYLE)
    add_values(ws, 1, 2 + len(headers), [target_model.name if target_model.name else target_model.id], HEADER_STYLE)
    add_values(ws, 2, 2 + len(headers), headers, HEADER_STYLE)

    ws.cell(row=3, column=1).value = "Matching"
    i = 4
    for s_r_id in sorted(r_id2r_id.iterkeys()):
        t_r_id = r_id2r_id[s_r_id]
        s = RED_STYLE if get_bounds(source_model.getReaction(s_r_id)) != get_bounds(target_model.getReaction(t_r_id)) \
            else BASIC_STYLE
        add_reaction(ws, s_r_id, source_model, i, 2, r2kegg, s, rev=True)
        add_reaction(ws, t_r_id, target_model, i, 2 + len(headers), r2kegg, s)
        i += 1

    ws.cell(row=i, column=1).value = "Similar in different compartments"
    i += 1
    for s_r_id in sorted(r_id2r_id_diff_comp.iterkeys()):
        add_reaction(ws, s_r_id, source_model, i, 2, r2kegg, rev=True)
        for t_r_id in r_id2r_id_diff_comp[s_r_id]:
            s = RED_STYLE if \
                get_bounds(source_model.getReaction(s_r_id)) != get_bounds(target_model.getReaction(t_r_id)) \
                else BASIC_STYLE
            add_reaction(ws, t_r_id, target_model, i, 2 + len(headers), r2kegg, s)
            i += 1

    ws.cell(row=i, column=1).value = "Unmapped"
    i += 1
    for s_r_id in sorted(s_r_ids_unmapped):
        add_reaction(ws, s_r_id, source_model, i, 2, r2kegg, rev=True)
        i += 1


def save_compartment_mappings(ws, source_model, target_model, c_id2c_id, c_ids_unmapped):
    headers = ["Name", "Id"]
    add_values(ws, 1, 2, [source_model.name if source_model.name else source_model.id], HEADER_STYLE)
    add_values(ws, 2, 2, reversed(headers), HEADER_STYLE)
    add_values(ws, 1, 2 + len(headers), [target_model.name if target_model.name else target_model.id], HEADER_STYLE)
    add_values(ws, 2, 2 + len(headers), headers, HEADER_STYLE)

    ws.cell(row=3, column=1).value = "Matching"
    i = 4
    for s_c_id in sorted(c_id2c_id.iterkeys()):
        t_c_id = c_id2c_id[s_c_id]
        add_compartment(ws, s_c_id, source_model, i, 2, rev=True)
        add_compartment(ws, t_c_id, target_model, i, 2 + len(headers))
        i += 1

    ws.cell(row=i, column=1).value = "Unmapped"
    i += 1
    for s_c_id in sorted(c_ids_unmapped):
        add_compartment(ws, s_c_id, source_model, i, 2, rev=True)
        i += 1


def add_metabolite(ws, m_id, model, row, col, m2kegg, m2chebi, rev=False):
    m = model.getSpecies(m_id)
    c_name = model.getCompartment(m.getCompartment()).getName()
    values = [format_m_name(m, model, False, False), m_id, c_name, m2chebi(m_id, model), m2kegg(m)]
    add_values(ws, row, col, reversed(values) if rev else values)


def add_reaction(ws, r_id, model, row, col, r2kegg, style=BASIC_STYLE, rev=False):
    r = model.getReaction(r_id)
    values = [get_sbml_r_formula(model, r), r.id, r.name] + list(get_bounds(r)) + [r2kegg(r)]
    add_values(ws, row, col, reversed(values) if rev else values, style)


def add_compartment(ws, c_id, model, row, col, rev=False):
    c = model.getCompartment(c_id)
    values = [c.name, c.id]
    add_values(ws, row, col, reversed(values) if rev else values)


def add_genes_to_reaction_list(sbml, path, out_path):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.get_active_sheet()

    worksheet.cell(row=1, column=7).value = "Gene association"
    worksheet.cell(row=1, column=8).value = "KEGG Reaction"
    input_doc = libsbml.SBMLReader().readSBML(sbml)
    model = input_doc.getModel()
    r_len = len(worksheet.rows)
    for i in xrange(2, r_len + 1):
        r_id = str(worksheet.cell(row=i, column=1).value)
        r = model.getReaction(r_id)
        worksheet.cell(row=i, column=7).value = get_gene_association(r)
        worksheet.cell(row=i, column=8).value = get_kegg_r_id(r)
    workbook.save(out_path)

