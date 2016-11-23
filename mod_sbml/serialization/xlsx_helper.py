import openpyxl
import openpyxl.styles
from openpyxl.styles import Style, Font
from openpyxl.styles.colors import Color

BASIC_STYLE = Style(font=Font(color=openpyxl.styles.colors.BLACK, sz=8))
RED_STYLE = Style(font=Font(color=openpyxl.styles.colors.RED, sz=8))
BLUE_STYLE = Style(font=Font(color=openpyxl.styles.colors.BLUE, sz=8))
GREEN_STYLE = Style(font=Font(color=openpyxl.styles.colors.GREEN, sz=8))
BOLD_STYLE = Style(font=Font(bold=True, sz=8))
HEADER_STYLE = BOLD_STYLE

__author__ = 'anna'


def serialize_values(x_name, xs, legend, y_lists, path):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(0)

    add_values(ws, 1, 1, [x_name], HEADER_STYLE)
    add_values(ws, 1, 2, xs)

    legend = iter(legend)
    j = 2
    for ys in y_lists:
        add_values(ws, j, 1, [next(legend)], HEADER_STYLE)
        add_values(ws, j, 2, ys)
        j += 1

    wb.save(path)


def get_values_by_column(path, i, skip_header=True):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.get_active_sheet()
    result = []
    for row in worksheet.rows:
        if skip_header:
            skip_header = False
        else:
            result.append(str(row[i].value))
    return result


def get_info(path, columns, start_row=1):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.get_active_sheet()
    r_len = len(worksheet.rows)
    for i in range(start_row, r_len + 1):
        yield [worksheet.cell(row=i, column=j).value for j in columns]


def add_values(ws, row, col, values, style=BASIC_STYLE):
    j = col
    for v in values:
        ws.cell(row=row, column=j).value = v
        ws.cell(row=row, column=j).style = style
        j += 1


def save_data(headers, data, wb=None, ws_name=None, ws_index=0, filename=None, styles=None):
    if not wb:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(ws_index, ws_name)
    add_values(ws, 1, 1, headers, HEADER_STYLE)
    row = 2
    if styles:
        styles = iter(styles)
    for values in data:
        add_values(ws, row, 1, values, style=next(styles) if styles else BASIC_STYLE)
        row += 1
    if filename:
        wb.save(filename)
    return ws





